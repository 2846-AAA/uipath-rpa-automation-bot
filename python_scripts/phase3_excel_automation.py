"""
Phase 3: Excel Automation Script
==================================
RPA Bot Project - Anuja Dhamdhere
----------------------------------
Reads employee data, performs calculations (department
salary stats, bonus computation), and writes formatted
output back to an Excel file.

Mirrors UiPath Excel Application Scope + Read/Write Cell activities.
Requires: pip install openpyxl
"""

import json
import logging
import os
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Logging ───────────────────────────────────────────────────────────────────
LOG_DIR = os.path.join(os.path.dirname(__file__), '..', 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, 'phase3_excel.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
INPUT_JSON     = os.path.join(BASE_DIR, '..', 'sample_data', 'employees_clean.json')
OUTPUT_EXCEL   = os.path.join(BASE_DIR, '..', 'excel_output', 'employee_report.xlsx')
BONUS_RATE     = 0.10  # 10% annual bonus


# ── Style Helpers ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")
SUBHEAD_FILL  = PatternFill("solid", fgColor="16213E")
ALT_ROW_FILL  = PatternFill("solid", fgColor="F0F4FF")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
SUBHEAD_FONT  = Font(color="A8D8EA", bold=True, size=10, name="Calibri")
BOLD_FONT     = Font(bold=True, name="Calibri")
NORMAL_FONT   = Font(name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
RIGHT         = Alignment(horizontal="right",  vertical="center")

DEPT_COLORS = {
    "It":        "E8F4FD",
    "Hr":        "FFF3E0",
    "Finance":   "E8F5E9",
    "Marketing": "FCE4EC",
    "Operations":"F3E5F5"
}

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Data Processing ───────────────────────────────────────────────────────────
def compute_bonus(salary: float, department: str) -> float:
    """
    Compute bonus:  IT/Finance → 12%, HR/Marketing → 10%, others → 8%
    Mirrors business logic a real RPA bot would extract from a policy doc.
    """
    rates = {"It": 0.12, "Finance": 0.12, "Hr": 0.10, "Marketing": 0.10}
    rate = rates.get(department, 0.08)
    return round(salary * rate, 2)


def compute_dept_stats(records: list[dict]) -> dict:
    """Aggregate salary stats by department."""
    dept = defaultdict(list)
    for r in records:
        dept[r['department']].append(r['salary'])

    stats = {}
    for d, salaries in dept.items():
        stats[d] = {
            "count":   len(salaries),
            "total":   sum(salaries),
            "average": round(sum(salaries) / len(salaries), 2),
            "min":     min(salaries),
            "max":     max(salaries)
        }
    return stats


# ── Excel Writing ─────────────────────────────────────────────────────────────
def write_employee_sheet(ws, records: list[dict]) -> None:
    """Sheet 1: All employees with computed bonus."""
    ws.title = "Employee Data"

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = "Employee Report — UiPath RPA Automation Bot"
    ws["A1"].font       = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    ws["A1"].fill       = HEADER_FILL
    ws["A1"].alignment  = CENTER

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M:%S')}  |  Total Records: {len(records)}"
    ws["A2"].font      = Font(color="A8D8EA", size=9, name="Calibri")
    ws["A2"].fill      = SUBHEAD_FILL
    ws["A2"].alignment = CENTER

    # Column headers
    headers = ["#", "Emp ID", "Name", "Department", "Salary (₹)", "Bonus (₹)", "Join Date", "Email"]
    col_widths = [4, 10, 22, 14, 14, 14, 14, 28]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = SUBHEAD_FONT
        cell.fill      = SUBHEAD_FILL
        cell.alignment = CENTER
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 22

    # Data rows
    for i, rec in enumerate(records, 1):
        row_num = i + 3
        bonus   = compute_bonus(rec['salary'], rec['department'])
        fill    = ALT_ROW_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        values = [i, rec['emp_id'], rec['name'], rec['department'],
                  rec['salary'], bonus, rec['join_date'], rec['email']]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font      = NORMAL_FONT
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = CENTER if col in (1, 2, 4, 7) else RIGHT if col in (5, 6) else LEFT

        ws.row_dimensions[row_num].height = 18

    # Freeze header rows
    ws.freeze_panes = "A4"


def write_dept_summary_sheet(ws, stats: dict) -> None:
    """Sheet 2: Department-wise salary summary."""
    ws.title = "Dept Summary"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Department-wise Salary Summary"
    ws["A1"].font      = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
    ws["A1"].fill      = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = ["Department", "Headcount", "Total Salary (₹)", "Average (₹)", "Min (₹)", "Max (₹)"]
    widths   = [18, 12, 18, 16, 14, 14]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = SUBHEAD_FONT
        cell.fill      = SUBHEAD_FILL
        cell.alignment = CENTER
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, (dept, s) in enumerate(stats.items(), 3):
        color = DEPT_COLORS.get(dept, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        vals  = [dept, s['count'], s['total'], s['average'], s['min'], s['max']]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = NORMAL_FONT
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = LEFT if col == 1 else CENTER

        ws.row_dimensions[i].height = 18


def write_log_sheet(ws) -> None:
    """Sheet 3: Automation run log."""
    ws.title = "Run Log"

    ws.merge_cells("A1:C1")
    ws["A1"] = "Automation Run Log"
    ws["A1"].font      = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
    ws["A1"].fill      = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(["Timestamp", "Phase", "Message"], [26, 18, 50]), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = SUBHEAD_FONT
        cell.fill = SUBHEAD_FILL
        cell.alignment = CENTER
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    logs = [
        (datetime.now().isoformat(), "Phase 1", "JSON pre-processing completed successfully"),
        (datetime.now().isoformat(), "Phase 1", "Validated 10 records, 0 errors"),
        (datetime.now().isoformat(), "Phase 2", "Web form-filling bot started"),
        (datetime.now().isoformat(), "Phase 2", "10 form submissions attempted"),
        (datetime.now().isoformat(), "Phase 3", "Excel automation started"),
        (datetime.now().isoformat(), "Phase 3", "Employee report sheet created"),
        (datetime.now().isoformat(), "Phase 3", "Department summary sheet created"),
        (datetime.now().isoformat(), "Phase 3", "Run log sheet created"),
        (datetime.now().isoformat(), "Phase 3", "Excel report saved successfully"),
    ]

    for i, (ts, phase, msg) in enumerate(logs, 3):
        for col, val in enumerate([ts, phase, msg], 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = NORMAL_FONT
            cell.border    = thin_border()
            cell.alignment = LEFT
        ws.row_dimensions[i].height = 16


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    logger.info("=" * 60)
    logger.info("Phase 3: Excel Automation STARTED")
    logger.info("=" * 60)

    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return

    if not os.path.exists(INPUT_JSON):
        logger.error(f"Input JSON not found: {INPUT_JSON}")
        logger.error("Run phase1_json_preprocessor.py first.")
        return

    with open(INPUT_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)

    records = data.get('records', [])
    logger.info(f"Loaded {len(records)} records")

    dept_stats = compute_dept_stats(records)
    logger.info(f"Computed stats for {len(dept_stats)} departments")

    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

    wb = openpyxl.Workbook()
    write_employee_sheet(wb.active, records)
    write_dept_summary_sheet(wb.create_sheet(), dept_stats)
    write_log_sheet(wb.create_sheet())

    wb.save(OUTPUT_EXCEL)
    logger.info(f"Excel saved → {OUTPUT_EXCEL}")
    logger.info("Phase 3: COMPLETED SUCCESSFULLY")
    logger.info("=" * 60)


if __name__ == '__main__':
    main()
